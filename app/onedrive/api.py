import requests
import pandas as pd
from io import BytesIO
from app.config import Config as cfg
from openpyxl import load_workbook
import os
import json
import hashlib
from datetime import datetime, date


relevant_columns = [
    "Job #",
    "Release #",
    "Job",
    "Description",
    "Fab Hrs",
    "Install HRS",
    "Paint color",
    "PM",
    "BY",
    "Released",
    "Fab Order",
    "Cut start",
    "Fitup comp",
    "Welded",
    "Paint Comp",
    "Ship",
    "Start install",
    "Comp. ETA",
    "Job Comp",
    "Invoiced",
    "Notes",
]


def get_access_token():
    """
    Get an access token for Microsoft Graph API using client credentials.
    """
    url = f"https://login.microsoftonline.com/{cfg.AZURE_TENANT_ID}/oauth2/v2.0/token"
    data = {
        "grant_type": "client_credentials",
        "client_id": cfg.AZURE_CLIENT_ID,
        "client_secret": cfg.AZURE_CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
    }
    r = requests.post(url, data=data)
    r.raise_for_status()
    return r.json()["access_token"]


def get_excel_dataframe():
    token = get_access_token()
    file_bytes = read_file_from_user_onedrive(
        token, cfg.ONEDRIVE_USER_EMAIL, cfg.ONEDRIVE_FILE_PATH
    )

    usecols = list(range(0, 20)) + [28]  # A-T (0-19) and AC (28)
    sheet_name = "Job Log"  

    df_all = pd.read_excel(
        BytesIO(file_bytes), header=2, usecols=usecols, sheet_name=sheet_name
    )

    # First data row is Excel row 4; store it before filtering
    df_all = df_all.reset_index(drop=False).rename(columns={"index": "_row0"})
    df_all["_excel_row"] = df_all["_row0"] + 4

    df_final = df_all.dropna(subset=["Job #", "Release #"]).copy()

    # Release # should be string (handle mixed types)
    df_final["Release #"] = df_final["Release #"].astype(str)

    wb = load_workbook(BytesIO(file_bytes), data_only=False)
    ws = wb[sheet_name]

    formula_col = 17  # Q
    formulas, has_formula = [], []

    for r in df_final["_excel_row"]:
        cell = ws.cell(row=int(r), column=formula_col)
        val = cell.value
        is_formula = isinstance(val, str) and val.startswith("=")
        formulas.append(val if is_formula else "")
        has_formula.append(is_formula)

    df_final["start_install_formula"] = formulas
    df_final["start_install_formulaTF"] = has_formula

    # Tidy up helper cols
    df_final = df_final.drop(columns=["_row0", "_excel_row"]).reset_index(drop=True)
    return df_final


def get_last_modified_time():
    access_token = get_access_token()
    file_path = cfg.ONEDRIVE_FILE_PATH  # e.g. "/Job Log 2.4 DM play toy (1).xlsm"
    user_email = cfg.ONEDRIVE_USER_EMAIL

    url = (
        f"https://graph.microsoft.com/v1.0/users/{user_email}/drive/root:/{file_path}:"
    )

    headers = {"Authorization": f"Bearer {access_token}"}

    r = requests.get(url, headers=headers)
    r.raise_for_status()
    data = r.json()

    return {
        "name": data.get("name"),
        "id": data.get("id"),
        "lastModifiedDateTime": data.get("lastModifiedDateTime"),
        "size": data.get("size"),
    }


def get_excel_data_with_timestamp():
    """
    Get the latest Excel data from OneDrive along with the last modified time.
    Returns a dict: {"last_modified": ..., "data": DataFrame}
    """
    last_modified_info = get_last_modified_time()
    df = get_excel_dataframe()
    return {
        "name": last_modified_info.get("name"),
        "last_modified_time": last_modified_info.get("lastModifiedDateTime"),
        "data": df,
    }


def read_file_from_user_onedrive(access_token, user_email, file_path):
    url = f"https://graph.microsoft.com/v1.0/users/{user_email}/drive/root:/{file_path}:/content"
    headers = {"Authorization": f"Bearer {access_token}"}
    r = requests.get(url, headers=headers)
    r.raise_for_status()
    return r.content


def list_root_contents(access_token, user_email):
    url = f"https://graph.microsoft.com/v1.0/users/{user_email}/drive/root/children"
    headers = {"Authorization": f"Bearer {access_token}"}
    r = requests.get(url, headers=headers)
    r.raise_for_status()
    return r.json()


def update_excel_cell(cell_address, value, worksheet_name="Job Log"):
    """
    Update a specific Excel cell via Microsoft Graph API

    Args:
        cell_address: Excel cell address (e.g., "O155")
        value: Value to set in the cell
        worksheet_name: Name of the worksheet (default: "Sheet1")

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        # Get access token
        access_token = get_access_token()

        # Build the URL for updating the cell
        user_email = cfg.ONEDRIVE_USER_EMAIL
        file_path = cfg.ONEDRIVE_FILE_PATH

        # URL to update a specific cell range
        url = f"https://graph.microsoft.com/v1.0/users/{user_email}/drive/root:/{file_path}:/workbook/worksheets/{worksheet_name}/range(address='{cell_address}')"

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json",
        }

        # Payload with the new value
        payload = {"values": [[value]]}

        print(f"Updating Excel cell {cell_address} with value: {value}")
        print(f"URL: {url}")

        # Make the PATCH request
        response = requests.patch(url, headers=headers, json=payload)

        if response.status_code == 200:
            print(f"Successfully updated cell {cell_address} with value '{value}'")
            return True
        else:
            print(f"Error updating Excel: {response.status_code} - {response.text}")
            return False

    except Exception as e:
        print(f"Exception in update_excel_cell: {e}")
        return False


# Excel Snapshot System
def capture_excel_snapshot(snapshot_date=None):
    """
    Capture a daily Excel snapshot for diff checking.
    
    Args:
        snapshot_date: Date for the snapshot (defaults to today)
        
    Returns:
        dict: Snapshot data with metadata
    """
    if snapshot_date is None:
        snapshot_date = date.today()
    
    try:
        # Get Excel data using existing function
        excel_data = get_excel_data_with_timestamp()
        
        if excel_data is None or "data" not in excel_data:
            raise Exception("Failed to download Excel data from OneDrive")
        
        df = excel_data["data"]
        
        # Create snapshot metadata
        snapshot_metadata = {
            "snapshot_date": snapshot_date.isoformat(),
            "captured_at": datetime.utcnow().isoformat(),
            "source_file": excel_data.get("name"),
            "last_modified": excel_data.get("last_modified_time"),
            "row_count": len(df),
            "columns": list(df.columns),
            "data_hash": hashlib.md5(df.to_string().encode()).hexdigest()
        }
        
        # Save snapshot
        snapshots_dir = "excel_snapshots"
        os.makedirs(snapshots_dir, exist_ok=True)
        
        snapshot_filename = f"snapshot_{snapshot_date.strftime('%Y%m%d')}"
        
        # Save DataFrame (pickle is most reliable for pandas)
        df.to_pickle(os.path.join(snapshots_dir, f"{snapshot_filename}.pkl"))
        
        # Save metadata
        with open(os.path.join(snapshots_dir, f"{snapshot_filename}_meta.json"), 'w') as f:
            json.dump(snapshot_metadata, f, indent=2)
        
        print(f"Snapshot captured: {snapshot_date} ({len(df)} rows)")
        
        return {
            "success": True,
            "snapshot_date": snapshot_date.isoformat(),
            "row_count": len(df),
            "file_path": os.path.join(snapshots_dir, f"{snapshot_filename}.pkl"),
            "metadata": snapshot_metadata
        }
        
    except Exception as e:
        print(f"Error capturing snapshot: {e}")
        return {
            "success": False,
            "error": str(e),
            "snapshot_date": snapshot_date.isoformat() if snapshot_date else None
        }


def capture_excel_snapshot_with_data(df, excel_data, snapshot_date=None):
    """
    Capture a snapshot using already downloaded Excel data.
    
    Args:
        df: DataFrame with Excel data
        excel_data: Excel data dict with metadata
        snapshot_date: Date for the snapshot (defaults to today)
        
    Returns:
        dict: Snapshot result
    """
    if snapshot_date is None:
        snapshot_date = date.today()
    
    try:
        # Create snapshot metadata
        snapshot_metadata = {
            "snapshot_date": snapshot_date.isoformat(),
            "captured_at": datetime.utcnow().isoformat(),
            "source_file": excel_data.get("name"),
            "last_modified": excel_data.get("last_modified_time"),
            "row_count": len(df),
            "columns": list(df.columns),
            "data_hash": hashlib.md5(df.to_string().encode()).hexdigest()
        }
        
        # Save snapshot
        snapshots_dir = "excel_snapshots"
        os.makedirs(snapshots_dir, exist_ok=True)
        
        snapshot_filename = f"snapshot_{snapshot_date.strftime('%Y%m%d')}"
        
        # Save DataFrame (pickle is most reliable for pandas)
        df.to_pickle(os.path.join(snapshots_dir, f"{snapshot_filename}.pkl"))
        
        # Save metadata
        with open(os.path.join(snapshots_dir, f"{snapshot_filename}_meta.json"), 'w') as f:
            json.dump(snapshot_metadata, f, indent=2)
        
        print(f"Snapshot captured: {snapshot_date} ({len(df)} rows)")
        
        return {
            "success": True,
            "snapshot_date": snapshot_date.isoformat(),
            "row_count": len(df),
            "file_path": os.path.join(snapshots_dir, f"{snapshot_filename}.pkl"),
            "metadata": snapshot_metadata
        }
        
    except Exception as e:
        print(f"Error capturing snapshot: {e}")
        return {
            "success": False,
            "error": str(e),
            "snapshot_date": snapshot_date.isoformat() if snapshot_date else None
        }


def load_snapshot(snapshot_date):
    """
    Load a snapshot for comparison.
    
    Args:
        snapshot_date: Date of snapshot to load
        
    Returns:
        tuple: (DataFrame, metadata) or (None, None) if not found
    """
    snapshot_filename = f"snapshot_{snapshot_date.strftime('%Y%m%d')}"
    snapshots_dir = "excel_snapshots"
    
    try:
        df = pd.read_pickle(os.path.join(snapshots_dir, f"{snapshot_filename}.pkl"))
        with open(os.path.join(snapshots_dir, f"{snapshot_filename}_meta.json"), 'r') as f:
            metadata = json.load(f)
        print(f"Snapshot loaded: {snapshot_date} ({len(df)} rows)")
        return df, metadata
    except FileNotFoundError:
        print(f"Snapshot not found for date: {snapshot_date}")
        return None, None
    except Exception as e:
        print(f"Error loading snapshot {snapshot_date}: {e}")
        return None, None


def get_latest_snapshot():
    """
    Get the most recent snapshot.
    
    Returns:
        tuple: (date, DataFrame, metadata) or (None, None, None) if no snapshots exist
    """
    snapshots_dir = "excel_snapshots"
    
    if not os.path.exists(snapshots_dir):
        return None, None, None
    
    snapshot_files = [f for f in os.listdir(snapshots_dir) if f.startswith("snapshot_") and f.endswith(".pkl")]
    
    if not snapshot_files:
        return None, None, None
    
    # Sort by filename (which includes date) to get the latest
    snapshot_files.sort(reverse=True)
    latest_file = snapshot_files[0]
    
    # Extract date from filename
    try:
        date_str = latest_file.replace("snapshot_", "").replace(".pkl", "")
        snapshot_date = datetime.strptime(date_str, "%Y%m%d").date()
        
        df, metadata = load_snapshot(snapshot_date)
        return snapshot_date, df, metadata
    except Exception as e:
        print(f"Error parsing date from snapshot filename {latest_file}: {e}")
        return None, None, None


def find_new_rows_in_excel(current_df, previous_df=None):
    """
    Find new rows by comparing current and previous DataFrames.
    
    Args:
        current_df: Current Excel data
        previous_df: Previous snapshot data (None if no previous data)
        
    Returns:
        DataFrame: New rows only
    """
    if previous_df is None:
        print("No previous snapshot found - treating all rows as new")
        return current_df.copy()
    
    # Create unique identifiers for comparison
    current_df = current_df.copy()
    previous_df = previous_df.copy()
    
    current_df['_identifier'] = current_df['Job #'].astype(str) + '-' + current_df['Release #'].astype(str)
    previous_df['_identifier'] = previous_df['Job #'].astype(str) + '-' + previous_df['Release #'].astype(str)
    
    # Find rows that exist in current but not in previous
    previous_identifiers = set(previous_df['_identifier'])
    new_rows = current_df[~current_df['_identifier'].isin(previous_identifiers)]
    
    # Clean up temporary column
    new_rows = new_rows.drop(columns=['_identifier'])
    
    print(f"Found {len(new_rows)} new rows out of {len(current_df)} total rows")
    return new_rows


def run_excel_snapshot_digest():
    """
    Run the Excel snapshot digest process:
    1. Capture current Excel data
    2. Compare with previous snapshot
    3. Save current data as new snapshot
    4. Return new rows found
    
    Returns:
        dict: Digest results with new rows and metadata
    """
    try:
        print("Starting Excel snapshot digest...")
        
        # Get current Excel data
        excel_data = get_excel_data_with_timestamp()
        if excel_data is None or "data" not in excel_data:
            raise Exception("Failed to download Excel data from OneDrive")
        
        current_df = excel_data["data"]
        print(f"Downloaded current Excel data: {len(current_df)} rows")
        
        # Get previous snapshot
        previous_date, previous_df, previous_metadata = get_latest_snapshot()
        
        if previous_df is not None:
            print(f"Found previous snapshot from {previous_date}: {len(previous_df)} rows")
        else:
            print("No previous snapshot found")
        
        # Find new rows
        new_rows = find_new_rows_in_excel(current_df, previous_df)
        
        # Save current data as new snapshot (reuse the downloaded data)
        snapshot_result = capture_excel_snapshot_with_data(current_df, excel_data)
        
        return {
            "success": True,
            "current_rows": len(current_df),
            "previous_rows": len(previous_df) if previous_df is not None else 0,
            "new_rows": len(new_rows),
            "new_rows_data": new_rows,
            "snapshot_captured": snapshot_result["success"],
            "previous_snapshot_date": previous_date.isoformat() if previous_date else None
        }
        
    except Exception as e:
        print(f"Excel snapshot digest failed: {e}")
        return {
            "success": False,
            "error": str(e)
        }
